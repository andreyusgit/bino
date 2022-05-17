from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
import math


positive = []  # создаем список со значениями х при которых зачет сдан
negative = []  # при этих значениях х зачет не сдан
averP = 0
averN = 0


# функция, которая импортирует все значения из таблицы в списки
def imp():
    global averP
    global averN
    data = load_workbook("data.xlsx")
    sheet = data['Sheet1']
    rows = sheet.max_row
    for index in range(2, rows):
        if sheet[f'B{index}'].value == 'да':
            averP += int(sheet[f'A{index}'].value)
            positive.append(sheet[f'A{index}'].value)
        elif sheet[f'B{index}'].value == 'нет':
            averN += int(sheet[f'A{index}'].value)
            negative.append(sheet[f'A{index}'].value)


# логистическая функция
def func(x):
    return (math.exp(-7 + x)) / (1 + math.exp(-7 + x))


# в этой функции происходит сравнение переданного в функцию значения и К.З.
def artificial_result(cv, x):
    if func(x) >= cv:
        return 'зачет'
    else:
        return 'незачет'


def click():
    time = int(Input.get())
    text = artificial_result(control_value, time)
    messagebox.showinfo(title='Результат', message=f'При таком времени подготовки, вероятнее всего, Вы получите {text}')


imp()
control_value = func(int(((averP / len(positive)) + (averN / len(negative))) / 2))
root = Tk()
ar = PhotoImage(file="arrow.gif")

root['bg'] = '#000000'
root.title('Предсказатель')
root.geometry('500x500')
root.resizable(width=False, height=False)

canvas = Canvas(root, width=500, height=500)
canvas.pack()

frame = Frame(root, bg='#010101')
frame.place(relx=0.025, rely=0.025, relwidth=0.95, relheight=0.95)
title = Label(frame, text='Введите сколько часов Вы готовились к зачету', bg='black', bd='5',
              relief='groove', font="Arial 13")
title.place(relx=0.15, rely=0.1)
arrow = Label(frame, image=ar)
arrow.place(relx=0.35, rely=0.25, relwidth=0.3, relheight=0.2)
Input = Entry(frame, fg='#010101', bg='white', font="Arial 25", justify='center')
Input.place(relx=0.195, rely=0.5)
butt = Button(frame, text='Ваш результат', font="Arial 30", justify='center', command=click)
butt.place(relx=0.25, rely=0.7)

root.mainloop()
