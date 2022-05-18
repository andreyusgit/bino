from openpyxl import load_workbook
import math
from array import *
import numpy as np
from numpy import linalg

positive = array('B', [])  # создаем массив со значениями х при которых зачет сдан
negative = array('B', [])  # при этих значениях х зачет не сдан
all_values = array('B', [])  # массив всех значений
all_results = array('B', [])  # массив всех результатов (1 - зачет, 0 - незачет)
aver = array('f', [0, 0])  # первое значение массива - среднее арифм среди положительных, второе - среди отрицательных
a = array('f', [0, 0])  # аргументы функции
P = np.array([])  # матрица P (нужна в методе наименьших квадратов)
Y = np.array([])  # матрица Y (нужна в методе наименьших квадратов)
W = np.array([])  # матрица весов W (нужна в методе наименьших квадратов)


# функция, которая импортирует все значения из таблицы в массивы
def imp():
    global aver
    sheet = load_workbook("data.xlsx")['Sheet1']
    rows = sheet.max_row

    for index in range(2, rows + 1):
        value = int(sheet[f'A{index}'].value)
        all_values.append(value)
        if sheet[f'B{index}'].value == 'да':
            aver[0] += value
            positive.append(value)
            all_results.append(1)
        elif sheet[f'B{index}'].value == 'нет':
            aver[1] += value
            negative.append(value)
            all_results.append(0)


# заполнение матриц Y, Р и W (при первом прогоне весов нет, поэтому матрица заполняется единицами)
def fill_matrix():
    len_of_arrays = len(all_values)
    global P
    P = np.array([[0.00] * 2 for i in range(len_of_arrays)])
    global Y
    Y = np.array([[0] * 1 for i in range(len_of_arrays)])
    global W
    W = np.array([[1.00] * len_of_arrays for i in range(len_of_arrays)])

    for i in range(len_of_arrays):
        for j in range(2):
            if j % 2 == 0:
                P[i][j] = 1
            else:
                P[i][j] = xs(all_values[i])

    for i in range(len_of_arrays):
        if all_results[i]:
            Y[i][0] = 1
        elif all_results[i]:
            Y[i][0] = 0


# расчет веса - чем дальше значение от эталонного, тем меньше его значимость
def weight():
    global W
    for i in range(len(positive) + len(negative)):
        W[i][i] = 1 - (math.fabs(func(all_results[i] * 30) - func(all_values[i])))
        # если результат незачет (0), то имеем выражение 0 * 30 = 0, результат зачет - 1 * 30 = 30


# в этой функции происходит сравнение переданного в функцию значения и К.З.
def artificial_result(cv, x):
    if func(x) >= cv:
        return 'зачет'
    else:
        return 'незачет'


# метод наименьших квадратов, используется для вычисления аргументов а0 и а1
def square_method(Y, P):
    temp = np.array(linalg.inv(((P.transpose()).dot(np.linalg.matrix_power(W, 2))).dot(P)))
    arguments = temp.dot(P.transpose().dot(np.linalg.matrix_power(W, 1)).dot(Y))

    a[0] = arguments[0]
    a[1] = arguments[1]


# замена для х
def xs(x):
    return math.exp(-x)


# функция по которой работает алгоритм
def func(x):
    return a[0] + a[1] * xs(x)


def main():
    imp()  # получаем все данные, по которым будем обучать алгоритм, из таблицы
    fill_matrix()  # заполняем матрицы значениями
    square_method(Y, P)  # первый раз ищем аргументы функции (без весов)
    weight()  # находим веса для каждого значения ищ нашей базы данных
    square_method(Y, P)  # второй прогон - с весами для каждого значения
    control_value = func(int(((aver[0] / len(positive)) + (aver[1] / len(negative))) / 2))  # находим КЗ
    return control_value
